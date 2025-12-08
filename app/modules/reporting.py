"""
Phase 14: Reporting Engine Module

Core business logic for analytics calculations, report generation,
and multi-format exports (CSV, PDF, Excel).
"""

from datetime import datetime, timedelta
from sqlalchemy import func, desc, and_
from flask import current_app
import csv
import io
import os


def calculate_revenue_metrics(start_date, end_date):
    """
    Calculate revenue metrics for a given period.
    
    Returns:
        dict with total_revenue, order_count, aov (average order value),
        refund_total, net_revenue, and comparison to previous period.
    """
    from app.models import Order, db
    
    # Current period
    current_orders = Order.query.filter(
        Order.status == 'paid',
        Order.created_at >= start_date,
        Order.created_at <= end_date
    ).all()
    
    total_revenue = sum(o.total_amount for o in current_orders) / 100  # Convert cents to dollars
    order_count = len(current_orders)
    aov = total_revenue / order_count if order_count > 0 else 0
    
    # Calculate previous period for comparison
    period_length = (end_date - start_date).days
    prev_start = start_date - timedelta(days=period_length)
    prev_end = start_date - timedelta(days=1)
    
    prev_orders = Order.query.filter(
        Order.status == 'paid',
        Order.created_at >= prev_start,
        Order.created_at <= prev_end
    ).all()
    
    prev_revenue = sum(o.total_amount for o in prev_orders) / 100
    prev_count = len(prev_orders)
    
    # Calculate growth percentages
    revenue_growth = ((total_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0
    order_growth = ((order_count - prev_count) / prev_count * 100) if prev_count > 0 else 0
    
    return {
        'total_revenue': round(total_revenue, 2),
        'order_count': order_count,
        'aov': round(aov, 2),
        'revenue_growth': round(revenue_growth, 1),
        'order_growth': round(order_growth, 1),
        'prev_revenue': round(prev_revenue, 2),
        'prev_order_count': prev_count,
        'period_start': start_date,
        'period_end': end_date
    }


def calculate_daily_revenue(start_date, end_date):
    """
    Get daily revenue breakdown for charting.
    
    Returns:
        list of dicts with date and revenue.
    """
    from app.models import Order, db
    
    results = db.session.query(
        func.date(Order.created_at).label('date'),
        func.sum(Order.total_amount).label('total')
    ).filter(
        Order.status == 'paid',
        Order.created_at >= start_date,
        Order.created_at <= end_date
    ).group_by(
        func.date(Order.created_at)
    ).order_by(
        func.date(Order.created_at)
    ).all()
    
    return [
        {'date': str(r.date), 'revenue': r.total / 100 if r.total else 0}
        for r in results
    ]


def calculate_product_performance(start_date, end_date, limit=20):
    """
    Calculate product performance metrics.
    
    Returns:
        list of products with revenue, units_sold, and avg_order_value.
    """
    from app.models import Order, OrderItem, Product, db
    
    results = db.session.query(
        Product.id,
        Product.name,
        func.sum(OrderItem.quantity).label('units_sold'),
        func.sum(OrderItem.price_at_purchase * OrderItem.quantity).label('total_revenue'),
        func.count(func.distinct(Order.id)).label('order_count')
    ).join(
        OrderItem, OrderItem.product_id == Product.id
    ).join(
        Order, Order.id == OrderItem.order_id
    ).filter(
        Order.status == 'paid',
        Order.created_at >= start_date,
        Order.created_at <= end_date
    ).group_by(
        Product.id, Product.name
    ).order_by(
        desc('total_revenue')
    ).limit(limit).all()
    
    return [
        {
            'product_id': r.id,
            'product_name': r.name,
            'units_sold': r.units_sold or 0,
            'total_revenue': (r.total_revenue or 0) / 100,
            'order_count': r.order_count or 0,
            'avg_order_value': ((r.total_revenue or 0) / r.order_count / 100) if r.order_count else 0
        }
        for r in results
    ]


def calculate_customer_clv(limit=100):
    """
    Calculate Customer Lifetime Value for top customers.
    
    Returns:
        list of customers with total_spent, order_count, avg_order_value, 
        first_purchase, last_purchase.
    """
    from app.models import Order, User, db
    
    results = db.session.query(
        User.id,
        User.email,
        User.username,
        func.count(Order.id).label('order_count'),
        func.sum(Order.total_amount).label('total_spent'),
        func.min(Order.created_at).label('first_purchase'),
        func.max(Order.created_at).label('last_purchase')
    ).join(
        Order, Order.user_id == User.id
    ).filter(
        Order.status == 'paid'
    ).group_by(
        User.id, User.email, User.username
    ).order_by(
        desc('total_spent')
    ).limit(limit).all()
    
    customers = []
    for r in results:
        total_spent = (r.total_spent or 0) / 100
        order_count = r.order_count or 0
        aov = total_spent / order_count if order_count > 0 else 0
        
        # Calculate customer tenure in days
        tenure_days = 0
        if r.first_purchase and r.last_purchase:
            tenure_days = (r.last_purchase - r.first_purchase).days
        
        customers.append({
            'user_id': r.id,
            'email': r.email,
            'username': r.username,
            'order_count': order_count,
            'total_spent': round(total_spent, 2),
            'avg_order_value': round(aov, 2),
            'first_purchase': r.first_purchase,
            'last_purchase': r.last_purchase,
            'tenure_days': tenure_days
        })
    
    return customers


def calculate_tax_report(start_date, end_date):
    """
    Calculate tax collected by jurisdiction for a period.
    
    Returns:
        dict with total_tax and breakdown by state/jurisdiction.
    """
    from app.models import Order, db
    import json
    
    # Get all paid orders in period
    orders = Order.query.filter(
        Order.status == 'paid',
        Order.created_at >= start_date,
        Order.created_at <= end_date
    ).all()
    
    # Parse shipping addresses to extract states
    tax_by_state = {}
    total_tax = 0
    
    for order in orders:
        # Assuming shipping_address contains state info
        # This is a simplified version - in production, you'd have proper tax records
        state = 'Unknown'
        if order.shipping_address:
            try:
                addr = json.loads(order.shipping_address) if isinstance(order.shipping_address, str) else order.shipping_address
                state = addr.get('state', 'Unknown')
            except (json.JSONDecodeError, TypeError):
                pass
        
        # Estimate tax at 8% (simplified - real implementation would use actual tax records)
        estimated_tax = order.total_amount * 0.08 / 100
        total_tax += estimated_tax
        
        if state not in tax_by_state:
            tax_by_state[state] = {'orders': 0, 'revenue': 0, 'tax': 0}
        
        tax_by_state[state]['orders'] += 1
        tax_by_state[state]['revenue'] += order.total_amount / 100
        tax_by_state[state]['tax'] += estimated_tax
    
    return {
        'total_tax': round(total_tax, 2),
        'total_orders': len(orders),
        'by_state': tax_by_state,
        'period_start': start_date,
        'period_end': end_date
    }


def calculate_traffic_metrics(start_date, end_date):
    """
    Calculate traffic and engagement metrics.
    
    Returns:
        dict with page_views, unique_visitors, bounce_rate, 
        avg_session_duration, top_pages, utm_sources.
    """
    from app.models import PageView, VisitorSession, db
    
    # Page views
    page_views = PageView.query.filter(
        PageView.timestamp >= start_date,
        PageView.timestamp <= end_date
    ).count()
    
    # Unique sessions
    unique_sessions = db.session.query(
        func.count(func.distinct(PageView.session_id))
    ).filter(
        PageView.timestamp >= start_date,
        PageView.timestamp <= end_date,
        PageView.session_id.isnot(None)
    ).scalar() or 0
    
    # Bounce rate from sessions
    sessions = VisitorSession.query.filter(
        VisitorSession.started_at >= start_date,
        VisitorSession.started_at <= end_date
    ).all()
    
    total_sessions = len(sessions)
    bounced_sessions = sum(1 for s in sessions if s.bounce)
    bounce_rate = (bounced_sessions / total_sessions * 100) if total_sessions > 0 else 0
    
    # Average session duration
    durations = [s.duration_seconds for s in sessions if s.duration_seconds]
    avg_duration = sum(durations) / len(durations) if durations else 0
    
    # Top pages
    top_pages = db.session.query(
        PageView.url,
        func.count(PageView.id).label('views')
    ).filter(
        PageView.timestamp >= start_date,
        PageView.timestamp <= end_date
    ).group_by(
        PageView.url
    ).order_by(
        desc('views')
    ).limit(10).all()
    
    # UTM source breakdown
    utm_sources = db.session.query(
        PageView.utm_source,
        func.count(PageView.id).label('views')
    ).filter(
        PageView.timestamp >= start_date,
        PageView.timestamp <= end_date,
        PageView.utm_source.isnot(None)
    ).group_by(
        PageView.utm_source
    ).order_by(
        desc('views')
    ).limit(10).all()
    
    # Device breakdown
    devices = db.session.query(
        PageView.device_type,
        func.count(PageView.id).label('views')
    ).filter(
        PageView.timestamp >= start_date,
        PageView.timestamp <= end_date
    ).group_by(
        PageView.device_type
    ).all()
    
    return {
        'page_views': page_views,
        'unique_sessions': unique_sessions,
        'bounce_rate': round(bounce_rate, 1),
        'avg_session_duration': int(avg_duration),
        'top_pages': [{'url': p.url, 'views': p.views} for p in top_pages],
        'utm_sources': [{'source': u.utm_source or 'Direct', 'views': u.views} for u in utm_sources],
        'devices': [{'type': d.device_type or 'Unknown', 'views': d.views} for d in devices],
        'period_start': start_date,
        'period_end': end_date
    }


def export_to_csv(data, headers, filename=None):
    """
    Generate CSV from report data.
    
    Args:
        data: list of dicts
        headers: list of column headers
        filename: optional filename for saving
    
    Returns:
        StringIO buffer with CSV data, or file path if filename provided.
    """
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(data)
    
    if filename:
        export_dir = os.path.join(current_app.static_folder, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        file_path = os.path.join(export_dir, filename)
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            f.write(output.getvalue())
        return file_path
    
    output.seek(0)
    return output


def export_to_excel(data, headers, filename, sheet_name='Report'):
    """
    Generate Excel file from report data.
    
    Args:
        data: list of dicts
        headers: list of column headers
        filename: filename for saving
        sheet_name: name of the Excel sheet
    
    Returns:
        file path of generated Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        current_app.logger.error("openpyxl not installed. Cannot generate Excel export.")
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4A90D9', end_color='4A90D9', fill_type='solid')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            # Format datetime objects
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=col, value=value)
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save file
    export_dir = os.path.join(current_app.static_folder, 'exports')
    os.makedirs(export_dir, exist_ok=True)
    file_path = os.path.join(export_dir, filename)
    wb.save(file_path)
    
    return file_path


def export_to_pdf(template_name, context, filename):
    """
    Generate PDF from template and data using WeasyPrint.
    
    Args:
        template_name: Jinja template name
        context: dict with template context
        filename: filename for saving
    
    Returns:
        file path of generated PDF.
    """
    try:
        from weasyprint import HTML
        from flask import render_template_string, render_template
    except ImportError:
        current_app.logger.error("WeasyPrint not installed. Cannot generate PDF export.")
        return None
    
    # Render template to HTML
    html_content = render_template(template_name, **context)
    
    # Generate PDF
    export_dir = os.path.join(current_app.static_folder, 'exports')
    os.makedirs(export_dir, exist_ok=True)
    file_path = os.path.join(export_dir, filename)
    
    HTML(string=html_content).write_pdf(file_path)
    
    return file_path


def execute_saved_report(report):
    """
    Execute a saved report configuration and return results.
    
    Args:
        report: SavedReport model instance
    
    Returns:
        dict with report data based on report_type.
    """
    config = report.config_json or {}
    
    # Parse date range from config
    end_date = datetime.utcnow()
    days = config.get('days', 30)
    start_date = end_date - timedelta(days=days)
    
    if report.report_type == 'revenue':
        return calculate_revenue_metrics(start_date, end_date)
    elif report.report_type == 'products':
        limit = config.get('limit', 20)
        return {'products': calculate_product_performance(start_date, end_date, limit)}
    elif report.report_type == 'customers':
        limit = config.get('limit', 100)
        return {'customers': calculate_customer_clv(limit)}
    elif report.report_type == 'traffic':
        return calculate_traffic_metrics(start_date, end_date)
    elif report.report_type == 'tax':
        return calculate_tax_report(start_date, end_date)
    else:
        return {'error': f'Unknown report type: {report.report_type}'}


def parse_user_agent(user_agent_string):
    """
    Parse user agent string to extract device, browser, and OS info.
    
    Returns:
        dict with device_type, browser, os.
    """
    if not user_agent_string:
        return {'device_type': 'unknown', 'browser': 'unknown', 'os': 'unknown'}
    
    ua = user_agent_string.lower()
    
    # Detect device type
    device_type = 'desktop'
    if 'mobile' in ua or 'android' in ua and 'mobile' in ua:
        device_type = 'mobile'
    elif 'tablet' in ua or 'ipad' in ua:
        device_type = 'tablet'
    
    # Detect browser
    browser = 'other'
    if 'chrome' in ua and 'edg' not in ua:
        browser = 'Chrome'
    elif 'firefox' in ua:
        browser = 'Firefox'
    elif 'safari' in ua and 'chrome' not in ua:
        browser = 'Safari'
    elif 'edg' in ua:
        browser = 'Edge'
    elif 'opera' in ua or 'opr' in ua:
        browser = 'Opera'
    
    # Detect OS
    os_name = 'other'
    if 'windows' in ua:
        os_name = 'Windows'
    elif 'mac' in ua or 'darwin' in ua:
        os_name = 'macOS'
    elif 'linux' in ua and 'android' not in ua:
        os_name = 'Linux'
    elif 'android' in ua:
        os_name = 'Android'
    elif 'iphone' in ua or 'ipad' in ua:
        os_name = 'iOS'
    
    return {
        'device_type': device_type,
        'browser': browser,
        'os': os_name
    }


def track_conversion(goal_id, session_id=None, user_id=None, value=None, 
                     source=None, medium=None, campaign=None):
    """
    Record a conversion event for a goal.
    
    Args:
        goal_id: ID of the ConversionGoal
        session_id: Session token (optional)
        user_id: User ID (optional)
        value: Conversion value (optional)
        source, medium, campaign: UTM params at time of conversion
    
    Returns:
        Conversion model instance or None if goal not found/inactive.
    """
    from app.models import ConversionGoal, Conversion, db
    
    goal = ConversionGoal.query.get(goal_id)
    if not goal or not goal.is_active:
        return None
    
    # Check if already converted in this session (if count_once_per_session is True)
    if goal.count_once_per_session and session_id:
        existing = Conversion.query.filter_by(
            goal_id=goal_id,
            session_id=session_id
        ).first()
        if existing:
            return existing
    
    conversion = Conversion(
        goal_id=goal_id,
        session_id=session_id,
        user_id=user_id,
        value=value or goal.target_value,
        source=source,
        medium=medium,
        campaign=campaign
    )
    
    db.session.add(conversion)
    db.session.commit()
    
    return conversion


def get_date_range_presets():
    """
    Get common date range presets for report filters.
    
    Returns:
        dict mapping preset names to (start_date, end_date) tuples.
    """
    now = datetime.utcnow()
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    return {
        'today': (today, now),
        'yesterday': (today - timedelta(days=1), today),
        'last_7_days': (today - timedelta(days=7), now),
        'last_30_days': (today - timedelta(days=30), now),
        'last_90_days': (today - timedelta(days=90), now),
        'this_month': (today.replace(day=1), now),
        'last_month': (
            (today.replace(day=1) - timedelta(days=1)).replace(day=1),
            today.replace(day=1) - timedelta(days=1)
        ),
        'this_year': (today.replace(month=1, day=1), now),
        'last_year': (
            today.replace(year=today.year-1, month=1, day=1),
            today.replace(month=1, day=1) - timedelta(days=1)
        )
    }
